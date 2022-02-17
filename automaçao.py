from bs4 import BeautifulSoup
from time import sleep
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from openpyxl import load_workbook
import PySimpleGUIQt as sg
from selenium.webdriver.chrome.service import Service as ChromeService
from subprocess import CREATE_NO_WINDOW



print('**************************************************************************')
print('************                 Derby                        ****************')
print('**************************************************************************')
print()
print('ATENÇAO CERTIFIQUE-SE DE QUE A PLANILHA QUE O ROBO VAI USAR ESTA FECHADA!')
print()

# --------listas----------
rodadas = []
time1 = []
time2 = []
lista_time1 = []
lista_time2 = []
lista_hora = []
placarT1 = []
placarT2 = []
lista_todos_os_jogos = [[], [], [], [], [], [], []]
lista_estrela = []
lista_e_estrela = [[], [], [], [], [], []]
lista_nao_estrela = [[], [], [], [], [], []]
lista_xpath = []
lista_pais = []
lista_ligas = []
lista_de_numeros_de_jogos = []
continuaçao = 0
# ------------------------
progresso = 0
comecar = 0
# ------------- layout ------------------------------------------
sg.theme('Dark2')
layout = [
    [sg.Text('{}'.format('           Derby'), font=('consolas', 20))],
    [sg.Output(size=(40, 3), font=('consolas', 12))],
    [sg.Button('começar', size=(20, 1)), sg.ProgressBar(15, orientation='h', size=(2, 0.8),bar_color=('green','gray'),key='progbar')]
]
estrela_confirmacao = [
    [sg.Text('você quer pegar as ligas estrelas?', font=('bolnd', 12))],
    [sg.Button('SIM', key='sim'), sg.Button('NÃO', key='nao')]
]
window = sg.Window('Derby investments', layout)
while True:
    event, values = window.read()
    if event == sg.WIN_CLOSED:
        break
    if continuaçao == 0:
        print('abrindo o excel')
        window['progbar'].update_bar(progresso + 1)
        # --------tabela excel----------
        tabela = load_workbook('PADRAO.xlsx')
        planilha = tabela.active
        # ------------------------------
        print('iniciando o chrome')
        window['progbar'].update_bar(progresso + 2)
        # --------inicializaçao webdriver----------
        options = Options()
        options.add_argument('--headless')
        chrome_service = ChromeService('chromedriver')
        chrome_service.creationflags = CREATE_NO_WINDOW
        options.add_experimental_option('excludeSwitches', ['enable-logging'])
        options.add_argument('window-size=400,800')
        navegador = webdriver.Chrome(options=options, service=chrome_service)
        navegador.get('https://www.flashscore.com.br/futebol')  # entrar no site

        # ------------------------------------------
        print('logando no site...')
        window['progbar'].update_bar(progresso + 3)
        print()
        # ---------------login no resultados------------------
        b_login = navegador.find_element(By.XPATH, '//*[@id="user-menu"]')  # achar o botao acessar
        navegador.execute_script("arguments[0].scrollIntoView();", b_login)  # apertar o botao acessar 1 de 2
        navegador.execute_script("arguments[0].click();", b_login)  # apertar o botao acessar 2 de 2
        sleep(3)
        navegador.find_element(By.XPATH, '//*[@id="email"]').send_keys('contapraseguir1@gmail.com')
        sleep(0.5)
        navegador.find_element(By.XPATH, '//*[@id="passwd"]').send_keys('Email.vp10')
        sleep(0.5)
        b_entrar = navegador.find_element(By.XPATH, '//*[@id="login"]')
        navegador.execute_script("arguments[0].scrollIntoView();", b_entrar)
        navegador.execute_script("arguments[0].click();", b_entrar)
        # ------------------------------------------------------
        window['progbar'].update_bar(progresso + 4)
        print('logado.')
        print()
        sleep(10)
        print('carregando coonteudo da primeira pagina')
        print()
        window['progbar'].update_bar(progresso + 5)
        # carregar todo conteudo da primeira pagina
        count = pausar = 0
        while True:
            count += 1
            try:
                b_mais_jogos = navegador.find_element(By.XPATH,
                                                      f'//*[@id="live-table"]/section/div/div/div[{count}]/span[2]')
            except:
                a = 0
            if b_mais_jogos.get_attribute('title') == "Exibir todos os jogos desta competição!":
                navegador.execute_script("arguments[0].scrollIntoView();", b_mais_jogos)
                navegador.execute_script("arguments[0].click();", b_mais_jogos)
                pausar = 0
            if b_mais_jogos.get_attribute('title') == "Ocultar todos os jogos desta competição!":
                pausar += 1
            if pausar == 100:
                break
        window['progbar'].update_bar(progresso + 6)
        print('conteudo da primeira pagina carregado com sucesso')
        print()
        # ---------------------------------------------------------------------------------------------

        conteudo_primeira_page = BeautifulSoup(navegador.page_source, 'html.parser')
        jogos_liga = conteudo_primeira_page.findAll('div', attrs={'class': 'sportName soccer'})
        cont = True  # cont = continuar
        p = 0
        contador = 0
        numero_de_ligas = 1
        numero_do_xpath = 0
        jogos_anteriores = 0
        numeros_de_jogos = 0
        liga_e_jogos = 0
        xpath_numero = 1
        index_recomeco = 0
        index = 0
        xpath = 0

        # ---------- conteudo de toda a pagina -----------
        conteudo_liga_e_jogos = str(jogos_liga)
        # ------------------------------------------------
        conteudo_liga_e_jogOs = conteudo_liga_e_jogos
        conteudo_liga_e_jogOS = conteudo_liga_e_jogos[43:]
        conteudo_liga_e_joGOS = conteudo_liga_e_jogos[43:]
        print('essa parte é a mais demorada tenha paciencia')
        print('organizando os dados...')
        print()
        window['progbar'].update_bar(progresso + 7)
        # ------------------------------------------------

        # ----------lipeza do nome do pais e liga---------
        conteudo_jogos__ = BeautifulSoup(conteudo_liga_e_jogos, 'html.parser')
        paises = conteudo_jogos__.findAll('span', attrs={'class': 'event__title--type'})
        ligas = conteudo_jogos__.findAll('span', attrs={'class': 'event__title--name'})
        for liga, pais in zip(ligas, paises):
            lista_ligas.append(liga.text)
            lista_pais.append(pais.text)
        # -----------------------------------------
        window['progbar'].update_bar(progresso + 8)
        # ----------lipeza do conteudo da estrela---------
        estrelas_ = conteudo_jogos__.findAll('div', attrs={'class': 'event__header'})
        for conteudo_topo_liga in estrelas_:
            conteudo_topo = str(conteudo_topo_liga)
            saber_se_e_estrela = conteudo_topo.count('eventStar--active')
            if saber_se_e_estrela == 0:
                lista_estrela.append('nao')
            else:
                lista_estrela.append('sim')
        # ------------------------------------------------
        window['progbar'].update_bar(progresso + 9)
        # --------------------xpath tabela da liga-------------------
        contador_ = 1
        while True:
            conti = conteudo_liga_e_jogOS.find('event__expanderBlock')
            if contador_ <= 1:
                contador_ = 2
                lista_xpath.append('//*[@id="live-table"]/section/div/div/div[1]/div[1]/div/span[2]')
            if conti != -1:
                inicio_index_liga = conteudo_liga_e_jogOS.find('event__expanderBlock')
                conteudo_pre_formatado_jogos_e_liga = conteudo_liga_e_jogOS[inicio_index_liga + 25:]
                fim_index_liga = conteudo_pre_formatado_jogos_e_liga.find('event__expanderBlock')
                liga_e_jogos = conteudo_pre_formatado_jogos_e_liga[:fim_index_liga]
                numeros_de_jogos = liga_e_jogos.count('event__participant event__participant--away')
                lista_de_numeros_de_jogos.append(numeros_de_jogos)
                conteudo_liga_e_jogOS = conteudo_pre_formatado_jogos_e_liga[fim_index_liga:]
                xpath_numero += numeros_de_jogos + 1
                xpath = f'//*[@id="live-table"]/section/div/div/div[{xpath_numero}]/div[1]/div/span[2]'
                lista_xpath.append(xpath)
            else:
                break
        window['progbar'].update_bar(progresso + 10)
        # -----------------------------------------------------------
        while cont:
            contador += 1
            print(contador)
            # ----------lipeza do nome do time 1---------
            inicio_index_nome_time1 = conteudo_liga_e_jogos.find('event__participant event__participant--home')
            fim_index_nome_time1 = conteudo_liga_e_jogos.find('event__logo event__logo--away')
            Time1 = conteudo_liga_e_jogos[inicio_index_nome_time1 + 45:fim_index_nome_time1 - 18]
            # tratamento de erro
            erro_ontbold = Time1.find('ontBold')
            erro_svg = Time1.find('<svg')
            erro_GOL = Time1.find('GOL')
            if erro_svg != -1:
                fim_index_nome_time1 = Time1.find('<svg')
                Time1 = Time1[:fim_index_nome_time1]

            if erro_GOL != -1:
                inicio_index_nome_time1 = Time1.find('">')
                fim_index_nome_time1 = Time1.find('<div')
                Time1 = Time1[inicio_index_nome_time1 + 2:fim_index_nome_time1]

            elif erro_ontbold != -1:
                inicio_index_nome_time1 = Time1.find('">')
                Time1 = Time1[inicio_index_nome_time1 + 2:]

            # -------------------------------------------

            # ----------lipeza do nome do time 2 ---------
            inicio_index_nome_time2 = conteudo_liga_e_jogos.find('event__participant event__participant--away')
            fim_index_nome_time2 = inicio_index_nome_time2 + 150
            Time2 = conteudo_liga_e_jogos[inicio_index_nome_time2 + 45:fim_index_nome_time2]
            fim_index_nome_time2 = Time2.find('</')
            Time2 = Time2[:fim_index_nome_time2]

            # tratamento de erro
            erro_ontbold = Time2.find('ontBold')
            erro_svg = Time2.find('<svg')
            erro_GOL = Time2.find('GOL')
            if erro_svg != -1:
                fim_index_nome_time2 = Time2.find('<svg')
                Time2 = Time2[:fim_index_nome_time2]

            if erro_GOL != -1:
                inicio_index_nome_time2 = Time2.find('">')
                fim_index_nome_time2 = Time2.find('<div')
                Time2 = Time2[inicio_index_nome_time2 + 2:fim_index_nome_time2]

            elif erro_ontbold != -1:
                inicio_index_nome_time2 = Time2.find('">')
                Time2 = Time2[inicio_index_nome_time2 + 2:]

            # --------------------------------------------

            # ----------condiçao para saber se o jogo ja rolou ---------
            fim = conteudo_liga_e_jogos.find('event__participant event__participant--away')
            conteudo_liga_e_jogoS = conteudo_liga_e_jogos[:fim + 300]
            hora_inicio = conteudo_liga_e_jogoS.find('event__time')
            if hora_inicio > 1:
                inicio_index_hora = conteudo_liga_e_jogoS.find('event__time')
                fim_index_hora = conteudo_liga_e_jogoS.find('</div><svg class="bet-ico')
                hora = conteudo_liga_e_jogoS[inicio_index_hora + 13:fim_index_hora]
                if len(hora) > 10:
                    hora = 'nao conseguimos obter o horario'
            else:
                hora = 'o jogo ja aconteceu ou nao vai mais acontecer'
            # -----------------------------------------------------------

            # ----------cortando a str em html por liga/jogo ---------
            inicio_index_corte = conteudo_liga_e_jogos.find('event__participant event__participant--away')
            conteudo_liga_e_jogos = conteudo_liga_e_jogos[inicio_index_corte + 200:]
            # --------------------------------------------
            lista_time1.append(Time1)
            lista_time2.append(Time2)
            lista_hora.append(hora)
            if inicio_index_corte == -1:
                break
        # ------------------------------------------------------------------------------------------------------------------
        window['progbar'].update_bar(progresso + 12)
        # -----------------------------------------

        print('pronto.')
        print()
        print('dividindo jogos estrela e nao estrela')
        print()
        window['progbar'].update_bar(progresso + 13)
        # ------------------------ traragem de jogos para dividir e igular todo os dados ---------------------
        numeros_de_jogos_ = 0
        for liga, pais, xpath__, jogos_, estrela_ in zip(lista_ligas, lista_pais, lista_xpath, lista_de_numeros_de_jogos,
                                                         lista_estrela):
            numeros_de_jogos_ += jogos_
            for i in range(jogos_):
                lista_todos_os_jogos[0].append(pais)
                lista_todos_os_jogos[1].append(liga)
                lista_todos_os_jogos[5].append(xpath__)
                lista_todos_os_jogos[6].append(estrela_)
        for time1_, time2_, hora_ in zip(lista_time1, lista_time2, lista_hora):
            lista_todos_os_jogos[2].append(time1_)
            lista_todos_os_jogos[3].append(time2_)
            lista_todos_os_jogos[4].append(hora_)
        for pais__, liga__, time1__, time2__, hora__, xpath___, estrela__ in zip(lista_todos_os_jogos[0],
                                                                                 lista_todos_os_jogos[1],
                                                                                 lista_todos_os_jogos[2],
                                                                                 lista_todos_os_jogos[3],
                                                                                 lista_todos_os_jogos[4],
                                                                                 lista_todos_os_jogos[5],
                                                                                 lista_todos_os_jogos[6]):
            if estrela__ == 'sim':
                lista_e_estrela[0].append(pais__)
                lista_e_estrela[1].append(liga__)
                lista_e_estrela[2].append(time1__)
                lista_e_estrela[3].append(time2__)
                lista_e_estrela[4].append(hora__)
                lista_e_estrela[5].append(xpath___)
            else:
                lista_nao_estrela[0].append(pais__)
                lista_nao_estrela[1].append(liga__)
                lista_nao_estrela[2].append(time1__)
                lista_nao_estrela[3].append(time2__)
                lista_nao_estrela[4].append(hora__)
                lista_nao_estrela[5].append(xpath___)
        # -------------------------------------------------------------------------------------------------
        window['progbar'].update_bar(progresso + 14)
        print('pronto.')
        print()
        conta = 0
        window['progbar'].update_bar(progresso + 15)
    while True:
        conta += 1
        window['progbar'].update_bar(progresso)
        if conta >= 2:
            estrela_confirmacao = [
                [sg.Text('você quer pegar as ligas estrelas?', font=('bolnd', 12))],
                [sg.Button('SIM', key='sim'), sg.Button('NÃO', key='nao')]
            ]
            navegador.get('https://www.flashscore.com.br/futebol')  # entrar no site
            tabela = load_workbook('PADRAO.xlsx')
            planilha = tabela.active
        window['progbar'].update_bar(progresso + 1)
        # --------------------- escolha tipo de liga estrela ou nao estrela--------------------------------
        estrela_ou_nao_estrelaa = sg.Window('Escolha os tipos de jogos', estrela_confirmacao, keep_on_top=True)
        estrela_ou_nao_estrela, nao_vou_usar = estrela_ou_nao_estrelaa.read()
        estrela_ou_nao_estrelaa.close()
        lista_escolhida = 0
        if estrela_ou_nao_estrela in 'sim':
            lista_escolhida = lista_e_estrela
        else:
            lista_escolhida = lista_nao_estrela
        window['progbar'].update_bar(progresso + 2)
        # -------------------------------------------------------------------------------------------------
        paiS___ = {}
        ligA___ = {}
        lista_paises = []
        lista_ligass = []
        # ----------------------------- escolher a liga ------------------------------
        count = 0
        for pais___ in lista_escolhida[0]:
            paiS___[pais___] = 0
        for paiss in paiS___:
            count += 1
            lista_paises.append(paiss)
        window['progbar'].update_bar(progresso + 3)
        escolha_pais = [
            [sg.Combo(lista_paises, key='box', size=(20, 0.7), font=('bonld', 10), )],
            [sg.Button('Continuar', font=('bold', 10))]]
        pais_escolhidoo = sg.Window('Escolha o pais', layout=escolha_pais, keep_on_top=True)
        event, values = pais_escolhidoo.read()
        pais_escolhidoo.close()
        pais_escolhido = values['box']
        count = 0
        for paisss, liga___ in zip(lista_escolhida[0], lista_escolhida[1]):
            if pais_escolhido == paisss:
                ligA___[liga___] = 0
        for ligaa in ligA___:
            count += 1
            lista_ligass.append(ligaa)
            window['progbar'].update_bar(progresso + 4)
        escolha_liga = [
            [sg.Combo(lista_ligass, key='box', size=(20, 0.7), font=('bold', 10), )],
            [sg.Button('Continuar', font=('bold', 10))]
        ]
        liga_escolhidaa = sg.Window('Escolha a liga', escolha_liga, keep_on_top=True)
        even, values = liga_escolhidaa.read()
        liga_escolhidaa.close()
        liga_escolhida = values['box']
        jogos_que_vou_usar = [[], [], [], [], [], []]
        for pais___, liga___, time1___, time2___, hora___, xpath____ in zip(lista_escolhida[0], lista_escolhida[1],
                                                                            lista_escolhida[2], lista_escolhida[3],
                                                                            lista_escolhida[4], lista_escolhida[5]):
            if pais_escolhido == pais___:
                if liga_escolhida == liga___:
                    jogos_que_vou_usar[0].append(pais___)
                    jogos_que_vou_usar[1].append(liga___)
                    jogos_que_vou_usar[2].append(time1___)
                    jogos_que_vou_usar[3].append(time2___)
                    jogos_que_vou_usar[4].append(hora___)
                    jogos_que_vou_usar[5].append(xpath____)
        window['progbar'].update_bar(progresso + 5)
        # ----------------------------------------------------------------------------------
        print('limpando dados antigos do excel')
        print()
        # --------------- limpar o lugar aonde vai colocar os jogos no excel ----------------
        contador = 1
        for celula in range(500):
            contador += 1
            planilha[f'A{contador}'] = ''
            planilha[f'B{contador}'] = ''
            planilha[f'C{contador}'] = ''
            planilha[f'D{contador}'] = ''
            planilha[f'E{contador}'] = ''
            planilha[f'F{contador}'] = ''
            planilha[f'G{contador}'] = ''
            planilha[f'H{contador}'] = ''
            planilha[f'I{contador}'] = ''
            planilha[f'J{contador}'] = ''
            planilha[f'K{contador}'] = ''
        # -----------------------------------------------------------------------------------
        window['progbar'].update_bar(progresso + 6)
        print('pronto.')
        print()
        print('adicionando os jogos de hj')
        print()
        # ----------------------- adiciona os jogos que vao acontecer ----------------------------
        linha = 1
        for _pais, _liga, _time1, _time2, _hora, _xpath in zip(jogos_que_vou_usar[0], jogos_que_vou_usar[1],
                                                               jogos_que_vou_usar[2], jogos_que_vou_usar[3],
                                                               jogos_que_vou_usar[4], jogos_que_vou_usar[5]):
            linha += 1
            planilha[f'H{linha}'] = _time1
            planilha[f'I{linha}'] = _time2
            planilha[f'J{linha}'] = _hora
            xpath = _xpath
        # -----------------------------------------------------------------------------------------
        window['progbar'].update_bar(progresso + 7)
        print('pronto')
        print()
        # ----------------------------- entrar na pagina de resultados dos jogos------------------------------
        sleep(5)
        b_liga_clasificacao = navegador.find_element(By.XPATH, f'{xpath}')
        sleep(0.5)
        navegador.execute_script("arguments[0].click();", b_liga_clasificacao)
        sleep(1)
        b_liga_resultados = navegador.find_element(By.XPATH, '//*[@id="li1"]')
        navegador.execute_script("arguments[0].click();", b_liga_resultados)
        sleep(5)
        # ------------------------------------------------------------------------------------------
        window['progbar'].update_bar(progresso + 8)
        print('caregando todos resultados da liga')
        print()
        b_ver_mais = 0
        try:
            b_ver_mais = navegador.find_element(By.XPATH, '//*[@id="live-table"]/div[1]/div/div/a')
        except:
            z = True
        # careega os dados da pagina
        c = True
        while c:
            sleep(5)
            try:
                navegador.execute_script("arguments[0].scrollIntoView();", b_ver_mais)
                navegador.execute_script("arguments[0].click();", b_ver_mais)
            except:
                c = False
        window['progbar'].update_bar(progresso + 9)
        print('pagina carregada')
        site = BeautifulSoup(navegador.page_source, 'html.parser')

        conteudo_time = site.prettify()
        conteudo_time = str(conteudo_time)
        fim = conteudo_time.find('¬~AA÷')
        conteudo_time = conteudo_time[fim + 20:]
        conteudo_time1 = conteudo_time.count('Rodada')
        print('jogos encontrado {}'.format(conteudo_time1))
        # ------------------------------------------------------------------------------
        window['progbar'].update_bar(progresso + 10)
        # --------------------- ajeitando jogos -----------------------------------------
        site = BeautifulSoup(navegador.page_source, 'html.parser')
        conteudo_dos_resultados = site.findAll('div', attrs={'id': 'tournament-page-data-results'})
        _nome_time1 = site.findAll('div', attrs={'class': 'event__participant--home'})
        placar_time1 = site.findAll('div', attrs={'class': 'event__score--home'})
        placar_time2 = site.findAll('div', attrs={'class': 'event__score--away'})
        _nome_time2 = site.findAll('div', attrs={'class': 'event__participant--away'})
        conteudo_dos_resultados = str(conteudo_dos_resultados)
        lista_rodadas = []
        lista_comparacao = []
        window['progbar'].update_bar(progresso + 11)
        while True:
            index_inicio_rodada = conteudo_dos_resultados.find('ER÷')
            index_fim_rodada = conteudo_dos_resultados.find('¬RW')
            inicio_bagui = conteudo_dos_resultados.find('CX÷')
            lista_comparacao.append(conteudo_dos_resultados[inicio_bagui + 3: index_fim_rodada])
            bagui = conteudo_dos_resultados[inicio_bagui:index_fim_rodada]
            if len(conteudo_dos_resultados[index_inicio_rodada + 10:index_fim_rodada]) < 10:
                if lista_comparacao.count(bagui) == 0:
                    lista_rodadas.append(conteudo_dos_resultados[index_inicio_rodada + 10:index_fim_rodada])

            lista_comparacao.append(bagui)
            conteudo_dos_resultados = conteudo_dos_resultados[index_fim_rodada + 5:]
            if index_inicio_rodada == -1:
                break
        # --------------------------------------------------------------------------------------
        window['progbar'].update_bar(progresso + 12)
        print('pronto')
        print()
        # ---------------------------pegar os nomes dos times-----------------------------------
        sleep(5)
        lista_times = []
        b_clasificacao = navegador.find_element(By.XPATH, '//*[@id="li3"]')
        navegador.execute_script("arguments[0].scrollIntoView();", b_clasificacao)
        navegador.execute_script("arguments[0].click();", b_clasificacao)
        conteudo_clasificacao = navegador.find_elements(By.CLASS_NAME, 'tableCellParticipant__name')
        for time in conteudo_clasificacao:
            lista_times.append(time.text)
        print('nome dos times pegado')
        print()
        print('adicionando os resultados no excel')
        window['progbar'].update_bar(progresso + 13)
        # --------------------------------------------------------------------------------------
        linha = 1
        for time in lista_times:
            linha += 1
            planilha[f'K{linha}'] = time
        linha = 1
        print('o nome dos times ja foram adicionados no excel')
        print()
        for rodada, time1, time2, placar1, placar2 in zip(lista_rodadas, _nome_time1, _nome_time2, placar_time1,
                                                          placar_time2):
            linha += 1
            planilha[f'A{linha}'] = rodada
            planilha[f'B{linha}'] = time1.text
            planilha[f'C{linha}'] = time2.text
            planilha[f'D{linha}'] = placar1.text
            planilha[f'E{linha}'] = placar2.text
            planilha[f'F{linha}'] = liga_escolhida
            planilha[f'G{linha}'] = pais_escolhido
        window['progbar'].update_bar(progresso + 14)
        tabela.save('PADRAO.xlsx')
        window['progbar'].update_bar(progresso + 15)
        print('planilha fechada voce ja pode abrir!')
        print()
        continuar = [
            [sg.Text('você quer pegar outra liga?', font=('bold', 12))],
            [sg.Button('SIM', key='sim'), sg.Button('NÃO', key='nao')]
        ]
        continuacao = sg.Window('CONTINUAR', continuar, keep_on_top=True)
        conti, values = continuacao.read()
        continuacao.close()
        if conti == 'nao':
            break
    window['progbar'].update_bar(progresso)
    continuaçao = 1
    event, values = window.read()
    if event == sg.WIN_CLOSED:
        navegador.close()
        break
